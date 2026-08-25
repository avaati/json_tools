#!/usr/bin/env python3
"""
flatten_json_to_excel.py

Flattens an arbitrarily-nested JSON array of objects (e.g. an Illumio API
export) into a flat Excel spreadsheet -- one row per top-level object.

Nesting rule:
    - Nested dict keys are joined to their parent with a colon:
        interfaces:network:href
    - A list of scalars, or a list of dicts, collapses into the SAME
      column(s) as a single-object would, with multiple values joined
      by "; " so each top-level object still produces exactly one row.
      e.g. two interfaces -> interfaces:name = "ens192; ens192"
    - Works to any depth, any number of objects -- no schema is hardcoded.

Usage:
    python flatten_json_to_excel.py input.json output.xlsx [--sep :] [--join "; "]
"""

import argparse
import json
import sys
from pathlib import Path

import pandas as pd


def flatten(obj, parent_key="", sep=":", join_str="; "):
    """
    Recursively flattens dicts/lists into a single-level dict.
    Returns {flat_key: scalar_value}.
    """
    items = {}

    if isinstance(obj, dict):
        if not obj:
            # empty dict -> blank cell for this key
            items[parent_key] = ""
            return items
        for k, v in obj.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else str(k)
            items.update(flatten(v, new_key, sep, join_str))
        return items

    if isinstance(obj, list):
        if not obj:
            items[parent_key] = ""
            return items
        # Flatten every element under the SAME parent_key, then merge
        # column-by-column across elements so the list still yields one row.
        sub_dicts = [flatten(el, parent_key, sep, join_str) for el in obj]
        all_keys = []
        seen = set()
        for d in sub_dicts:
            for k in d.keys():
                if k not in seen:
                    seen.add(k)
                    all_keys.append(k)
        for key in all_keys:
            values = [
                str(d[key]) for d in sub_dicts
                if key in d and d[key] not in (None, "")
            ]
            items[key] = join_str.join(values)
        return items

    # scalar (str, int, float, bool, None)
    items[parent_key] = "" if obj is None else obj
    return items


def flatten_records(records, sep=":", join_str="; "):
    return [flatten(rec, "", sep, join_str) for rec in records]


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_json", help="Path to input JSON file (array of objects)")
    parser.add_argument("output_xlsx", help="Path to write the output .xlsx file")
    parser.add_argument("--sep", default=":", help="Separator for nested keys (default ':')")
    parser.add_argument("--join", dest="join_str", default="; ",
                         help="Delimiter used to join multiple list values in one cell (default '; ')")
    parser.add_argument("--sheet-name", default="Workloads", help="Excel sheet name")
    args = parser.parse_args()

    data = json.loads(Path(args.input_json).read_text())
    if isinstance(data, dict):
        # allow a single object, or a dict with a top-level list under some key
        data = [data]
    if not isinstance(data, list):
        sys.exit("Input JSON must be an array of objects (or a single object).")

    records = flatten_records(data, sep=args.sep, join_str=args.join_str)
    df = pd.DataFrame(records)

    df.to_excel(args.output_xlsx, sheet_name=args.sheet_name, index=False)
    _format_workbook(args.output_xlsx, args.sheet_name)
    print(f"Wrote {len(df)} rows x {len(df.columns)} columns -> {args.output_xlsx}")


def _format_workbook(path, sheet_name):
    """Light formatting: bold header, frozen top row, autofilter, sane column widths."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    ws = wb[sheet_name]

    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")

    for row in ws.iter_rows():
        for cell in row:
            cell.font = header_font if cell.row == 1 else body_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 60)

    wb.save(path)


if __name__ == "__main__":
    main()
